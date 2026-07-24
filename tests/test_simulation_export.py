from __future__ import annotations

import unittest

from openpyxl import load_workbook

from backend.app.simulation_export import build_simulation_workbook


class SimulationExportTests(unittest.TestCase):
    def test_workbook_contains_design_static_modal_and_optimization_results(self) -> None:
        payload = {
            "name": "best-028",
            "generated_at": "2026-07-24T10:00:00Z",
            "design_intent": {
                "part_type": "bushing",
                "geometry": {
                    "outer_diameter_mm": 76,
                    "inner_diameter_mm": 28.5,
                    "inner_core_length_mm": 71,
                    "outer_core_length_mm": 39.5,
                },
                "material": {"name": "rubber", "shore_a": 55},
            },
            "client_targets": {
                "target_kx_n_mm": 88.4,
                "target_ky_n_mm": 294.5,
                "target_kz_n_mm": 294.5,
            },
            "best_design": {
                "case_id": "BEST-028",
                "source": "trained static-FEM neural surrogate",
                "withinTolerance": True,
                "kx": 85.1,
                "ky": 301.0,
                "kz": 300.8,
                "maxRelativeError": 0.037,
                "rmsRelativeError": 0.029,
                "geometry": {"outer_diameter_mm": 76, "inner_diameter_mm": 28.5},
            },
            "static_stiffness": {
                "material": "rubber",
                "youngs_modulus_mpa": 1.1,
                "directions": [
                    {
                        "engineering_axis": "x",
                        "mesh_axis": "Z",
                        "displacement_mm": 1,
                        "reaction_force_n": 81.35,
                        "stiffness_n_per_mm": 81.35,
                    },
                    {
                        "engineering_axis": "y",
                        "mesh_axis": "X",
                        "displacement_mm": 1,
                        "reaction_force_n": 303.09,
                        "stiffness_n_per_mm": 303.09,
                    },
                ],
            },
            "modal_fem": {
                "mode": 1,
                "num_modes": 2,
                "material": "rubber",
                "fundamental_hz": 274,
                "modes": [
                    {"mode_number": 1, "frequency_hz": 274, "eigenvalue": 2.96e6},
                    {"mode_number": 2, "frequency_hz": 1200, "eigenvalue": 5.68e7},
                ],
                "pca": {
                    "mode_scores": [
                        {"mode_number": 1, "scores": [0.5, -0.2]},
                        {"mode_number": 2, "scores": [-0.5, 0.2]},
                    ],
                    "components": [
                        {
                            "component": 1,
                            "explained_variance_ratio": 0.62,
                            "cumulative_variance_ratio": 0.62,
                            "characteristic": "Primary mode-family variation",
                            "dominant_mode": 1,
                            "dominant_frequency_hz": 274,
                            "dominant_axis": "z",
                            "axis_energy": {"x": 0.2, "y": 0.3, "z": 0.5},
                            "score_min": -0.5,
                            "score_max": 0.5,
                        }
                    ],
                },
            },
            "mesh_summary": {
                "status": "ok",
                "nodes": 3024,
                "hexahedra": 2304,
                "surface_mesh": {"faces": [{"points": [1, 2, 3]}]},
            },
            "design_space_cases": [
                {
                    "case_id": "RB-001",
                    "geometry": {
                        "outer_diameter_mm": 76,
                        "inner_diameter_mm": 24,
                        "inner_core_length_mm": 40,
                    },
                }
            ],
        }

        workbook = load_workbook(build_simulation_workbook(payload), data_only=True)

        self.assertIn("Summary", workbook.sheetnames)
        self.assertIn("Design Parameters", workbook.sheetnames)
        self.assertIn("Static Stiffness", workbook.sheetnames)
        self.assertIn("Modal Results", workbook.sheetnames)
        self.assertIn("Mode PCA", workbook.sheetnames)
        self.assertIn("Mesh", workbook.sheetnames)
        self.assertIn("Optimization", workbook.sheetnames)
        self.assertIn("Design Space", workbook.sheetnames)
        self.assertEqual(workbook["Static Stiffness"]["F2"].value, 81.35)
        self.assertEqual(workbook["Modal Results"]["B3"].value, 1200)
        self.assertEqual(workbook["Optimization"]["B2"].value, "BEST-028")
        mesh_values = [cell.value for cell in workbook["Mesh"]["A"]]
        self.assertFalse(any("Surface Mesh" in str(value) for value in mesh_values))


if __name__ == "__main__":
    unittest.main()
