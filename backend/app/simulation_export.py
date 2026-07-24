"""Create compact Excel workbooks for Resonance AI simulation results."""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
import json
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


BRAND = "10243F"
CAD = "178078"
HEADER_FILL = PatternFill("solid", fgColor=BRAND)
SECTION_FILL = PatternFill("solid", fgColor="EAF1F8")
GOOD_FILL = PatternFill("solid", fgColor="E9F7EF")
WARNING_FILL = PatternFill("solid", fgColor="FFF7D6")
BAD_FILL = PatternFill("solid", fgColor="FFF0EF")
THIN_BORDER = Border(bottom=Side(style="thin", color="D5DFEA"))


def build_simulation_workbook(payload: dict[str, Any]) -> BytesIO:
    """Build an in-memory XLSX report from compact frontend result data."""

    if not isinstance(payload, dict):
        raise ValueError("Simulation report payload must be a JSON object.")

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _write_summary(summary, payload)
    _write_design_parameters(workbook, payload)
    _write_static_stiffness(workbook, payload)
    _write_modal_results(workbook, payload)
    _write_mode_pca(workbook, payload)
    _write_mapping_sheet(workbook, "Mesh", payload.get("mesh_summary"))
    _write_optimization(workbook, payload)
    _write_design_cases(workbook, payload)
    _write_mapping_sheet(workbook, "Shape PCA", payload.get("shape_pca"))
    _write_mapping_sheet(workbook, "Analytical", payload.get("analytical_estimate"))

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _write_summary(sheet: Worksheet, payload: dict[str, Any]) -> None:
    best = _mapping(payload.get("best_design"))
    static = _mapping(payload.get("static_stiffness"))
    modal = _mapping(payload.get("modal_fem"))
    intent = _mapping(payload.get("design_intent"))
    material = _mapping(intent.get("material"))

    sheet.append(["Resonance AI simulation report", ""])
    sheet.merge_cells("A1:B1")
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    rows = [
        ("Generated (UTC)", payload.get("generated_at") or datetime.now(UTC).isoformat()),
        ("Model", payload.get("name") or "model"),
        ("Part type", intent.get("part_type") or ""),
        ("Material", material.get("name") or static.get("material") or modal.get("material") or ""),
        ("Best case", best.get("case_id") or ""),
        ("Best-design source", best.get("source") or ""),
        ("Target match", _target_match_label(best)),
        ("Static stiffness available", "Yes" if static else "No"),
        ("Modal FEM available", "Yes" if modal else "No"),
        ("Solved modes", modal.get("num_modes") or len(_list(modal.get("modes")))),
        ("Selected contour mode", modal.get("mode") or ""),
        ("Fundamental frequency (Hz)", modal.get("fundamental_hz") or ""),
        ("Notes", "POC output. Validate material, boundary conditions, mesh convergence, and test correlation before engineering release."),
    ]
    for label, value in rows:
        sheet.append([label, _safe_cell(value)])
    _style_key_value_sheet(sheet)
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 86
    sheet.freeze_panes = "A2"


def _write_design_parameters(workbook: Workbook, payload: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Design Parameters")
    sheet.append(["Section", "Parameter", "Value", "Unit"])
    intent = _mapping(payload.get("design_intent"))
    sections = (
        ("Geometry", intent.get("geometry")),
        ("Material", intent.get("material")),
        ("Simulation settings", payload.get("simulation_settings")),
        ("Client targets", payload.get("client_targets")),
        ("Global mesh template", payload.get("global_mesh_template")),
    )
    for section, values in sections:
        for key, value in _flatten_mapping(_mapping(values)):
            label, unit = _parameter_label_and_unit(key)
            sheet.append([section, label, _safe_cell(value), unit])
    _style_table(sheet, freeze="A2", auto_filter=True)


def _write_static_stiffness(workbook: Workbook, payload: dict[str, Any]) -> None:
    data = _mapping(payload.get("static_stiffness"))
    directions = _list(data.get("directions"))
    if not data and not directions:
        return

    sheet = workbook.create_sheet("Static Stiffness")
    headers = [
        "Client axis",
        "Client target (N/mm)",
        "Solver axis",
        "Displacement (mm)",
        "Reaction force (N)",
        "Stiffness (N/mm)",
        "Absolute error (N/mm)",
        "Error (%)",
    ]
    sheet.append(headers)
    targets = _mapping(payload.get("client_targets"))
    calibration = _mapping(data.get("calibration"))
    reference_targets = _mapping(calibration.get("reference_targets_n_per_mm"))
    for item_value in directions:
        item = _mapping(item_value)
        axis = str(item.get("engineering_axis") or "").lower()
        target = _number(
            reference_targets.get(f"k{axis}"),
            _number(targets.get(f"target_k{axis}_n_mm"), _number(targets.get(f"k{axis}"))),
        )
        stiffness = _number(item.get("stiffness_n_per_mm"))
        absolute_error = abs(stiffness - target) if target is not None and stiffness is not None else None
        error_percent = absolute_error / abs(target) * 100 if absolute_error is not None and target else None
        sheet.append(
            [
                f"K{axis}",
                target,
                item.get("mesh_axis"),
                _number(item.get("displacement_mm")),
                _number(item.get("reaction_force_n")),
                stiffness,
                absolute_error,
                error_percent,
            ]
        )

    _style_table(sheet, freeze="A2", auto_filter=True)
    for row in range(2, sheet.max_row + 1):
        for column in (2, 4, 5, 6, 7):
            sheet.cell(row, column).number_format = "0.00"
        sheet.cell(row, 8).number_format = '0.00"%"'
    if sheet.max_row >= 2:
        error_range = f"H2:H{sheet.max_row}"
        sheet.conditional_formatting.add(
            error_range,
            CellIsRule(operator="lessThan", formula=["5"], fill=GOOD_FILL),
        )
        sheet.conditional_formatting.add(
            error_range,
            CellIsRule(operator="between", formula=["5", "10"], fill=WARNING_FILL),
        )
        sheet.conditional_formatting.add(
            error_range,
            CellIsRule(operator="greaterThan", formula=["10"], fill=BAD_FILL),
        )

    detail_start = sheet.max_row + 3
    sheet.cell(detail_start, 1, "Model details")
    sheet.cell(detail_start, 1).font = Font(bold=True, color=BRAND)
    details = [
        ("Material", data.get("material")),
        ("Centerline axis", data.get("centerline_axis")),
        ("Fixed interface", data.get("fixed_interface")),
        ("Effective modulus (MPa)", data.get("youngs_modulus_mpa")),
        ("Poisson ratio", data.get("poisson_ratio")),
        ("Inner interface nodes", data.get("inner_node_count")),
        ("Outer interface nodes", data.get("outer_node_count")),
        ("Limitations", data.get("model_limitations")),
    ]
    for offset, (label, value) in enumerate(details, start=1):
        sheet.cell(detail_start + offset, 1, label)
        sheet.cell(detail_start + offset, 2, _safe_cell(value))
    _autosize(sheet)


def _write_modal_results(workbook: Workbook, payload: dict[str, Any]) -> None:
    data = _mapping(payload.get("modal_fem"))
    modes = _list(data.get("modes"))
    if not modes:
        return

    pca = _mapping(data.get("pca"))
    score_by_mode = {
        int(_number(item.get("mode_number"), 0) or 0): _list(item.get("scores"))
        for item in map(_mapping, _list(pca.get("mode_scores")))
    }
    component_count = max((len(scores) for scores in score_by_mode.values()), default=0)
    headers = [
        "Mode",
        "Frequency (Hz)",
        "Eigenvalue",
        "Selected contour",
        "Material",
        *[f"PC{index + 1} score" for index in range(component_count)],
    ]
    sheet = workbook.create_sheet("Modal Results")
    sheet.append(headers)
    for mode_value in modes:
        mode = _mapping(mode_value)
        mode_number = int(_number(mode.get("mode_number"), 0) or 0)
        scores = score_by_mode.get(mode_number, [])
        sheet.append(
            [
                mode_number,
                _number(mode.get("frequency_hz")),
                _number(mode.get("eigenvalue")),
                "Yes" if mode_number == int(_number(data.get("mode"), 0) or 0) else "No",
                data.get("material") or "generic",
                *[_number(scores[index]) if index < len(scores) else None for index in range(component_count)],
            ]
        )
    _style_table(sheet, freeze="A2", auto_filter=True)
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 2).number_format = "0.000"
        sheet.cell(row, 3).number_format = "0.000E+00"


def _write_mode_pca(workbook: Workbook, payload: dict[str, Any]) -> None:
    modal = _mapping(payload.get("modal_fem"))
    pca = _mapping(modal.get("pca"))
    components = _list(pca.get("components"))
    if not components:
        return

    sheet = workbook.create_sheet("Mode PCA")
    sheet.append(
        [
            "Component",
            "Variance (%)",
            "Cumulative (%)",
            "Characteristic",
            "Dominant mode",
            "Dominant frequency (Hz)",
            "Dominant axis",
            "X energy (%)",
            "Y energy (%)",
            "Z energy (%)",
            "Score min",
            "Score max",
        ]
    )
    for component_value in components:
        component = _mapping(component_value)
        energy = _mapping(component.get("axis_energy"))
        sheet.append(
            [
                f"PC{component.get('component')}",
                _percent_value(component.get("explained_variance_ratio")),
                _percent_value(component.get("cumulative_variance_ratio")),
                component.get("characteristic"),
                component.get("dominant_mode"),
                _number(component.get("dominant_frequency_hz")),
                str(component.get("dominant_axis") or "").upper(),
                _percent_value(energy.get("x")),
                _percent_value(energy.get("y")),
                _percent_value(energy.get("z")),
                _number(component.get("score_min")),
                _number(component.get("score_max")),
            ]
        )
    _style_table(sheet, freeze="A2", auto_filter=True)
    for row in range(2, sheet.max_row + 1):
        for column in (2, 3, 8, 9, 10):
            sheet.cell(row, column).number_format = '0.0"%"'


def _write_optimization(workbook: Workbook, payload: dict[str, Any]) -> None:
    best = _mapping(payload.get("best_design"))
    if not best:
        return
    sheet = workbook.create_sheet("Optimization")
    sheet.append(["Parameter", "Value", "Unit"])
    rows = [
        ("Case ID", best.get("case_id"), ""),
        ("Source", best.get("source"), ""),
        ("Inside tolerance", "Yes" if best.get("withinTolerance") else "No", ""),
        ("Predicted Kx", best.get("kx"), "N/mm"),
        ("Predicted Ky", best.get("ky"), "N/mm"),
        ("Predicted Kz", best.get("kz"), "N/mm"),
        ("Maximum target error", _percent_value(best.get("maxRelativeError")), "%"),
        ("RMS target error", _percent_value(best.get("rmsRelativeError")), "%"),
    ]
    for label, value, unit in rows:
        sheet.append([label, _safe_cell(value), unit])
    for key, value in _flatten_mapping(_mapping(best.get("geometry"))):
        label, unit = _parameter_label_and_unit(key)
        sheet.append([label, _safe_cell(value), unit])
    _style_table(sheet, freeze="A2", auto_filter=False)


def _write_design_cases(workbook: Workbook, payload: dict[str, Any]) -> None:
    cases = [_mapping(item) for item in _list(payload.get("design_space_cases"))[:5000]]
    if not cases:
        return
    geometry_keys = sorted(
        {
            key
            for item in cases
            for key in _mapping(item.get("geometry")).keys()
            if not isinstance(_mapping(item.get("geometry")).get(key), (dict, list))
        }
    )
    sheet = workbook.create_sheet("Design Space")
    sheet.append(["Case ID", *[_parameter_label_and_unit(key)[0] for key in geometry_keys]])
    for item in cases:
        geometry = _mapping(item.get("geometry"))
        sheet.append([item.get("case_id"), *[_safe_cell(geometry.get(key)) for key in geometry_keys]])
    _style_table(sheet, freeze="A2", auto_filter=True)


def _write_mapping_sheet(workbook: Workbook, title: str, value: Any) -> None:
    mapping = _mapping(value)
    if not mapping:
        return
    sheet = workbook.create_sheet(title[:31])
    sheet.append(["Parameter", "Value", "Unit"])
    for key, item in _flatten_mapping(mapping):
        label, unit = _parameter_label_and_unit(key)
        sheet.append([label, _safe_cell(item), unit])
    _style_table(sheet, freeze="A2", auto_filter=False)


def _style_key_value_sheet(sheet: Worksheet) -> None:
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 1).font = Font(bold=True, color=BRAND)
        sheet.cell(row, 1).fill = SECTION_FILL
        for column in (1, 2):
            sheet.cell(row, column).border = THIN_BORDER
            sheet.cell(row, column).alignment = Alignment(vertical="top", wrap_text=True)


def _style_table(sheet: Worksheet, *, freeze: str, auto_filter: bool) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 25
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = freeze
    if auto_filter and sheet.max_row >= 1:
        sheet.auto_filter.ref = sheet.dimensions
    _autosize(sheet)


def _autosize(sheet: Worksheet) -> None:
    for column_index in range(1, sheet.max_column + 1):
        longest = 0
        for row_index in range(1, min(sheet.max_row, 300) + 1):
            value = sheet.cell(row_index, column_index).value
            longest = max(longest, len(str(value or "")))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(longest + 2, 12), 48)


def _flatten_mapping(
    mapping: dict[str, Any],
    prefix: str = "",
    *,
    excluded: Iterable[str] = ("surface_mesh", "fem_mesh", "contour_png_base64", "faces", "points"),
) -> list[tuple[str, Any]]:
    excluded_keys = set(excluded)
    flattened: list[tuple[str, Any]] = []
    for key, value in mapping.items():
        if key in excluded_keys:
            continue
        path = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            flattened.extend(_flatten_mapping(value, path, excluded=excluded_keys))
        elif isinstance(value, list):
            if len(value) <= 12 and all(not isinstance(item, (dict, list)) for item in value):
                flattened.append((path, ", ".join(str(item) for item in value)))
            else:
                flattened.append((path, f"{len(value)} item(s)"))
        else:
            flattened.append((path, value))
    return flattened


def _parameter_label_and_unit(key: str) -> tuple[str, str]:
    final_key = key.split(".")[-1]
    units = (
        ("_kg_m3", "kg/m3"),
        ("_n_per_mm", "N/mm"),
        ("_mm3", "mm3"),
        ("_mm2", "mm2"),
        ("_mm", "mm"),
        ("_mpa", "MPa"),
        ("_hz", "Hz"),
        ("_deg", "deg"),
    )
    unit = ""
    clean_key = final_key
    for suffix, label in units:
        if clean_key.endswith(suffix):
            clean_key = clean_key[: -len(suffix)]
            unit = label
            break
    label = clean_key.replace("_", " ").strip().title()
    if "." in key:
        parent = key.rsplit(".", 1)[0].replace(".", " / ").replace("_", " ").title()
        label = f"{parent} / {label}"
    return label, unit


def _safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=True, sort_keys=True)
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    return text[:32767]


def _target_match_label(best: dict[str, Any]) -> str:
    if not best:
        return ""
    return "Inside tolerance" if best.get("withinTolerance") else "Closest candidate; outside tolerance"


def _mapping(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def _number(value: Any, fallback: float | None = None) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return fallback
    return number


def _percent_value(value: Any) -> float | None:
    number = _number(value)
    return number * 100 if number is not None else None
